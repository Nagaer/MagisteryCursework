from openpyxl import load_workbook


def strfication(k, things):
    if k == 1:
        return (things + 'один')
    elif k <= 5:
        return (things + 'несколько')
    elif k <= 10:
        return (things + 'много')
    else:
        return (things + 'очень много')


def drunkification(k):
    if k < 5:
        return ('Трезвый')
    elif k <= 10:
        return ('Слабо нетрезвый')
    elif k <= 15:
        return ('Средне нетрезвый')
    else:
        return ('Сильно нетрезвый')


def updatedict(ditc, tipe, key):
    if ditc.get(tipe) is None:
        ditc[tipe] = {}
    if type(key) == list:
        for x in key:
            if ditc[tipe].get(x) is None:
                ditc[tipe][x] = 1
            else:
                ditc[tipe][x] += 1
    else:
        if ditc[tipe].get(key) is None:
            ditc[tipe][key] = 1
        else:
            ditc[tipe][key] += 1


dictDiffRating = {'Проезжая часть - Гололедица': 40.0,
                  'Проезжая часть - Загрязненное': 37.0,
                  'Проезжая часть - Заснеженное': 38.0,
                  'Проезжая часть - Мокрая дорога': 37.0,
                  'Проезжая часть - Обработанное противогололедными материалами': 37.0,
                  'Проезжая часть - Пыльное': 36.6,
                  'Проезжая часть - Со снежным накатом': 37.8,
                  'Проезжая часть - Сухая дорога': 36.6,

                  'Изменения. Режим движения - Движение полностью перекрыто': 42.0,
                  'Изменения. Режим движения - Движение частично перекрыто': 39.3,
                  'Изменения. Режим движения - Режим движения не изменялся': 36.6,

                  'ТС: много': 42.0,
                  'ТС: несколько': 39.5,
                  'ТС: один': 37.6,

                  'Освещение - В темное время суток, освещение включено': 37.0,
                  'Освещение - В темное время суток, освещение не включено': 40.0,
                  'Освещение - В темное время суток, освещение отсутствует': 42.0,
                  'Освещение - Светлое время суток': 36.6,
                  'Освещение - Сумерки': 36.8,

                  'Использовался ли ремень - Да': 36.6,
                  'Использовался ли ремень - Нет': 42.0,

                  'Оставление места ДТП - Не установлен': 36.6,
                  'Оставление места ДТП - Нет (не скрывался)': 36.6,
                  'Оставление места ДТП - Осталось на месте ДТП': 36.6,
                  'Оставление места ДТП - Скрылся и впоследствии не установлен': 42.0,
                  'Оставление места ДТП - Скрылся, впоследствии разыскан (установлен)': 40.0,

                  'Степень опьянения - Трезвый': 36.6,
                  'Степень опьянения - Слабо нетрезвый': 38.0,
                  'Степень опьянения - Средне нетрезвый': 40.0,
                  'Степень опьянения - Сильно нетрезвый': 42.0,

                  'Тяжесть последствия - Не пострадал': 36.6,
                  'Тяжесть последствия - Получил травмы с оказанием разовой медицинской помощи, к категории раненый не относится': 37.0,
                  'Тяжесть последствия - Раненый, находящийся (находившийся) на амбулаторном лечении, либо в условиях дневного стационара': 39.0,
                  'Тяжесть последствия - Раненый, находящийся (находившийся) на стационарном лечении': 40.0,
                  'Тяжесть последствия - Скончался после прибытия в больницу': 42.0,
                  'Тяжесть последствия - Скончался на месте ДТП до приезда скорой медицинской помощи': 42.0,
                  'Тяжесть последствия - Скончался на месте ДТП по прибытию скорой медицинской помощи, но до транспортировки в мед. организацию': 42.0,
                  'Тяжесть последствия - Скончался при транспортировке': 42.0,

                  'Тяжесть последствия - Не пострадал': 36.6,
                  'Тяжесть последствия - Получил травмы с оказанием разовой медицинской помощи, к категории раненый не относится': 37.0,
                  'Тяжесть последствия - Раненый, находящийся (находившийся) на амбулаторном лечении, либо в условиях дневного стационара': 39.0,
                  'Тяжесть последствия - Раненый, находящийся (находившийся) на стационарном лечении': 40.0,
                  'Тяжесть последствия - Скончался после прибытия в больницу': 42.0,
                  'Тяжесть последствия - Скончался на месте ДТП до приезда скорой медицинской помощи': 42.0,
                  'Тяжесть последствия - Скончался на месте ДТП по прибытию скорой медицинской помощи, но до транспортировки в мед. организацию': 42.0,
                  'Тяжесть последствия - Скончался при транспортировке': 42.0,

                  'ТСОД': 40.0,
                  'ТСОД - Не установлены': 36.6,

                  'Технические неисправности': 40.0,
                  'Технические неисправности - Технические неисправности отсутствуют': 36.6,

                  'Погибших: один': 40.5,
                  'Погибших: несколько': 41.5,
                  'Погибших: много': 42.0,

                  'Пострадавших: один': 37.5,
                  'Пострадавших: несколько': 39.5,
                  'Пострадавших: много': 40.0,

                  'Участников: один': 36.6,
                  'Участников: несколько': 38.0,
                  'Участников: много': 39.0,
                  'Участников: очень много': 41.0}


class Extractor(object):
    def __init__(self):
        self.numberDTP = 0

        self.totalNumTS = 0
        self.totalNumMembers = 0
        self.totalNumDeaths = 0
        self.totalNumHurts = 0

        self.dictResult = {}

    def extract(self, number):
        for qqq in range(1, number+1):
            wb = load_workbook('./Карточки ДТП, часть ' + str(qqq) + '.xlsx')
            print('Подключение к Базе Данных №' + str(qqq) + ' завершено...')

            for x in wb.sheetnames:
                # print('Обрабатываю ', x)
                sheet = wb[x]
                self.numberDTP += 1
                for i in range(1, sheet.max_row):
                    s = sheet.cell(row=i, column=1).value
                    if s == 'Вид\xa0ДТП':  # Ошибки нет, не трогать
                        TypeDTP = 'Вид ДТП - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Недостатки транспортно-эксплуатационного содержания улично-дорожной сети:':
                        TSOD = []
                        k = 0
                        while True:
                            TSOD.append('ТСОД - ' + sheet.cell(row=i + k, column=2).value)
                            if sheet.cell(row=i + k + 1, column=1).value is None:
                                k += 1
                            else:
                                break
                    elif s == 'Состояние проезжей части:':
                        if sheet.cell(row=i, column=2).value == 'Сухое':
                            Road = 'Проезжая часть - Сухая дорога'
                        elif sheet.cell(row=i, column=2).value == 'Мокрое':
                            Road = 'Проезжая часть - Мокрая дорога'
                        else:
                            Road = 'Проезжая часть - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Фокторы, оказывающие влияние на режим движения:':
                        if sheet.cell(row=i, column=2).value == 'Сведения отсутствуют':
                            Factor = 'Факторов, оказывающих влияние на режим движения, не обнаружено'
                        else:
                            Factor = sheet.cell(row=i, column=2).value
                    elif s == 'Освещение:':
                        Lighting = 'Освещение - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Изменения в режиме движения:':
                        Mode = 'Изменения. Режим движения - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Количество ТС':
                        NumTS = int(sheet.cell(row=i, column=2).value)
                        self.totalNumTS += NumTS
                        NumTS = strfication(NumTS, 'ТС: ')

                        NumMembers = int(sheet.cell(row=i, column=4).value)
                        self.totalNumMembers += NumMembers
                        NumMembers = strfication(NumMembers, 'Участников: ')

                        NumDeaths = int(sheet.cell(row=i, column=6).value)
                        self.totalNumDeaths += NumDeaths
                        NumDeaths = strfication(NumDeaths, 'Погибших: ')

                        NumHurts = int(sheet.cell(row=i, column=8).value)
                        self.totalNumHurts += NumHurts
                        NumHurts = strfication(NumHurts, 'Пострадавших: ')
                    elif s == 'Сведения об оставлении места ДТП':
                        Leaving = 'Оставление места ДТП - ' + sheet.cell(row=i, column=2).value
                    elif (s == 'Категория участника') and ((sheet.cell(row=i, column=2).value == 'Водитель') or (
                            sheet.cell(row=i, column=2).value == 'Пассажир')):
                        Belt = 'Использовался ли ремень - ' + sheet.cell(row=i, column=6).value
                    elif s == 'Степень тяжести последствий':
                        if 'Скончался в течение' in sheet.cell(row=i, column=2).value:
                            Severity = 'Тяжесть последствия - Скончался после прибытия в больницу'
                        else:
                            Severity = 'Тяжесть последствия - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Технические неисправности':
                        Issues = 'Технические неисправности - ' + sheet.cell(row=i, column=2).value
                    elif s == 'Степень опьянения':
                        if sheet.cell(row=i, column=2).value == '':
                            Drunk = 'Степень трезвости - Трезвый'
                        else:
                            Drunk = 'Степень трезвости - ' + drunkification(int(sheet.cell(row=i, column=2).value))

                # Общая статистика
                # updatedict(dictResult, 'Вид ДТП', TypeDTP)
                updatedict(self.dictResult, 'Число участников', NumMembers)
                updatedict(self.dictResult, 'Число погибших', NumDeaths)
                updatedict(self.dictResult, "Число раненых", NumHurts)

                # Водитель
                updatedict(self.dictResult, "Ремень", Belt)
                updatedict(self.dictResult, "Сведения об оставлении места ДТП", Leaving)
                updatedict(self.dictResult, "Степень тяжести последствий", Severity)
                updatedict(self.dictResult, 'Степень трезвости', Drunk)

                # Автомобиль
                updatedict(self.dictResult, 'Количество ТС', NumTS)
                updatedict(self.dictResult, 'Технические неисправности', Issues)

                # Дорога
                updatedict(self.dictResult, "Дорога", Road)
                updatedict(self.dictResult, "Изменения в режиме движения", Mode)

                # Среда
                updatedict(self.dictResult, "Освещение", Lighting)
                updatedict(self.dictResult, "ТСОД", TSOD)
                updatedict(self.dictResult, 'Факторы, оказывающие влияние на режим движения', Factor)

        keys = list(self.dictResult.keys())
        keys.sort()
        dictKeys = {}
        dictStrResult = {}
        numKeys = 0
        for key in keys:
            print('=====' + key + '======')
            dictKeys[key] = list(self.dictResult[key].keys())
            dictKeys[key].sort()
            dictStrResult[key] = {}
            for in_key in dictKeys[key]:
                dictStrResult[key][in_key] = str(self.dictResult[key][in_key] / self.numberDTP * 100) + '%'
                numKeys += 1
                print(in_key + ' = ' + dictStrResult[key][in_key])

        riskTemp = 0.0

        keysV = ["Ремень", "Сведения об оставлении места ДТП", "Степень тяжести последствий", "Степень трезвости"]
        riskTempV = 0.0

        keysA = ["Количество ТС", "Технические неисправности", "Число участников", "Число погибших", "Число раненых"]
        riskTempA = 0.0

        keysD = ["Дорога", "Изменения в режиме движения"]
        riskTempD = 0.0

        keysS = ["Освещение", "ТСОД", "Факторы, оказывающие влияние на режим движения"]
        riskTempS = 0.0

        for key in keys:
            for in_key in dictKeys[key]:
                if dictDiffRating.get(in_key):
                    riskIndicator = dictDiffRating[in_key]
                elif dictDiffRating.get(key):
                    riskIndicator = dictDiffRating[key]
                else:
                    riskIndicator = 39.3
                riskTemp += self.dictResult[key][in_key] / self.numberDTP * riskIndicator
                if key in keysV:
                    riskTempV += self.dictResult[key][in_key] / self.numberDTP * riskIndicator
                if key in keysA:
                    riskTempA += self.dictResult[key][in_key] / self.numberDTP * riskIndicator
                if key in keysD:
                    riskTempD += self.dictResult[key][in_key] / self.numberDTP * riskIndicator
                if key in keysS:
                    riskTempS += self.dictResult[key][in_key] / self.numberDTP * riskIndicator

        riskTemp /= (len(keysV) + len(keysA) + len(keysD) + len(keysS))
        riskTempV /= len(keysV)
        riskTempA /= len(keysA)
        riskTempD /= len(keysD)
        riskTempS /= len(keysS)

        print('Общая рисковая температура: ' + str(riskTemp))
        print('Рисковая температура категории "Водитель": ' + str(riskTempV))
        print('Рисковая температура категории Автомобиль": ' + str(riskTempA))
        print('Рисковая температура категории "Дорога": ' + str(riskTempD))
        print('Рисковая температура категории "Среда": ' + str(riskTempS))

        return [riskTempV, riskTempA, riskTempD, riskTempS]
